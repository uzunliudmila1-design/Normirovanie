import os
import sys
import json
import base64
import sqlite3
import subprocess
import tempfile
from datetime import datetime
from flask import Flask, request, jsonify, render_template, g, send_file
from anthropic import Anthropic
from dotenv import load_dotenv
try:
    import openpyxl
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False

CLAUDE_BIN = "/Users/liudmila/Library/Application Support/Claude/claude-code/2.1.87/claude.app/Contents/MacOS/claude"

# Гарантируем UTF-8 для stdout/stderr (Python 3.9 на macOS может использовать ASCII)
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")

load_dotenv()

# ─── База оборудования завода ─────────────────────────────────────────────────
EQUIPMENT_DB_PATH = os.path.join(os.path.dirname(__file__), "Данные", "Обороудование.xlsx")
_EQUIPMENT_TEXT = None   # кешируем результат форматирования

# ─── Лог анализа ──────────────────────────────────────────────────────────────
import threading
_analysis_logs = []          # [{ts, level, source, message}]
_analysis_logs_lock = threading.Lock()
_ANALYSIS_LOG_MAX = 500


def alog(source, message, level="info"):
    """Добавляет запись в лог анализа."""
    entry = {
        "ts": datetime.now().strftime("%H:%M:%S"),
        "level": level,
        "source": source,
        "message": message
    }
    with _analysis_logs_lock:
        _analysis_logs.append(entry)
        if len(_analysis_logs) > _ANALYSIS_LOG_MAX:
            _analysis_logs[:] = _analysis_logs[-_ANALYSIS_LOG_MAX:]
    print(f"[LOG:{level}] [{source}] {message}", flush=True)


# ─── Каталог изделий ──────────────────────────────────────────────────────────
PRODUCTS_BASE_PATH = os.path.join(os.path.dirname(__file__), "Изделия")


def _safe_products_path(*parts):
    """Собирает путь внутри PRODUCTS_BASE_PATH, защита от path traversal."""
    full = os.path.realpath(os.path.join(PRODUCTS_BASE_PATH, *parts))
    base = os.path.realpath(PRODUCTS_BASE_PATH)
    if not full.startswith(base + os.sep) and full != base:
        return None
    return full


def load_equipment_text():
    """Загружает Данные/Обороудование.xlsx и возвращает форматированный текст для промпта."""
    global _EQUIPMENT_TEXT
    if _EQUIPMENT_TEXT is not None:
        return _EQUIPMENT_TEXT

    if not _OPENPYXL_AVAILABLE or not os.path.exists(EQUIPMENT_DB_PATH):
        _EQUIPMENT_TEXT = ""
        return _EQUIPMENT_TEXT

    wb = openpyxl.load_workbook(EQUIPMENT_DB_PATH, read_only=True, data_only=True)
    ws = wb['Лист1']

    # Группируем по операциям
    by_op = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[1]:
            continue
        name = str(row[1]).replace('\xa0', ' ').strip()
        цех = str(row[2]).strip() if row[2] else '—'
        ops = [str(x).strip() for x in [row[4], row[5], row[6]] if x]
        for op in ops:
            by_op.setdefault(op, []).append((name, цех))

    wb.close()

    lines = ["БАЗА ОБОРУДОВАНИЯ ЗАВОДА (используй только это оборудование при нормировании):"]
    lines.append("Формат: • Наименование [Цех/подразделение] — для каждой операции\n")
    for op in sorted(by_op.keys()):
        items = by_op[op]
        lines.append(f"{op}:")
        seen = set()
        for name, цех in items:
            key = (name[:50], цех)
            if key in seen:
                continue
            seen.add(key)
            lines.append(f"  • {name[:80]} [{цех}]")

    _EQUIPMENT_TEXT = "\n".join(lines)
    return _EQUIPMENT_TEXT
# ─────────────────────────────────────────────────────────────────────────────

# ─── Типовые маршруты ────────────────────────────────────────────────────────
TYPICAL_ROUTES_PATH = os.path.join(
    os.path.dirname(__file__), "Маршруты изделий", "Типовые маршруты v1.xlsx"
)
_TYPICAL_ROUTES_TEXT = None


def load_typical_routes():
    """Загружает Типовые маршруты v1.xlsx и возвращает форматированный текст для промпта."""
    global _TYPICAL_ROUTES_TEXT
    if _TYPICAL_ROUTES_TEXT is not None:
        return _TYPICAL_ROUTES_TEXT

    if not _OPENPYXL_AVAILABLE or not os.path.exists(TYPICAL_ROUTES_PATH):
        _TYPICAL_ROUTES_TEXT = ""
        return _TYPICAL_ROUTES_TEXT

    wb = openpyxl.load_workbook(TYPICAL_ROUTES_PATH, read_only=True, data_only=True)
    ws = wb.active
    lines = ["КАТАЛОГ ТИПОВЫХ МАРШРУТОВ ЗАВОДА (236 маршрутов):"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        num, route = row[0], row[1]
        if num and route:
            lines.append(f"  {num}: {route}")
    wb.close()
    _TYPICAL_ROUTES_TEXT = "\n".join(lines)
    return _TYPICAL_ROUTES_TEXT
# ─────────────────────────────────────────────────────────────────────────────

app = Flask(__name__)
app.config["JSON_AS_ASCII"] = False
DB_PATH = os.path.join(os.path.dirname(__file__), "norming.db")

SYSTEM_PROMPT = """Ты — система автоматического нормирования времени производственных операций на машиностроительном заводе.

РЕЖИМ РАБОТЫ определяется набором предоставленных документов:

РЕЖИМ А — Чертёж + Маршрутная карта:
  Тебе предоставлены чертёж и МК. Нормируй операции из МК, используя чертёж для расчёта параметров.

РЕЖИМ Б — Только чертёж (без маршрутной карты):
  Тебе предоставлен только чертёж и каталог типовых маршрутов завода.
  Алгоритм работы:
  1. Проанализируй чертёж: тип детали, материал, геометрия, размеры, точность, шероховатость
  2. По характеристикам детали выбери НАИБОЛЕЕ ПОДХОДЯЩИЙ типовой маршрут из каталога
  3. Укажи выбранный маршрут (его номер и операции) в поле "обоснование" первой операции
  4. Пронормируй каждую операцию выбранного маршрута как обычно
  Нумерацию операций присваивай сам: 010, 020, 030...

Тебе могут быть предоставлены:
1. PDF-файл чертежа детали — содержит геометрию, размеры, допуски, материал, шероховатость (ОБЯЗАТЕЛЬНО)
2. PDF-файл маршрутной карты — содержит перечень технологических операций, оборудование, инструмент (необязательно)
3. Размер партии (количество деталей) — используется для расчёта t_пз на деталь
4. (Опционально) PDF типового техпроцесса — эталонный перечень операций для данного семейства изделий. Если предоставлен: используй его как основу перечня операций

Твоя задача:
- Проанализировать предоставленные документы
- Для каждой операции САМОСТОЯТЕЛЬНО рассчитать нормы времени
- Вернуть результат строго в формате JSON (без пояснений вне JSON)

КРИТИЧЕСКИ ВАЖНО — РАСЧЁТ ВРЕМЕНИ:
- ЗАПРЕЩЕНО использовать времена из маршрутной карты как основу для нормы
- Маршрутная карта нужна ТОЛЬКО для: перечня операций, их последовательности, указанного оборудования и инструментов
- РАССЧИТЫВАЙ t_шт полностью самостоятельно по методике нормирования машиностроения:
  t_шт = t_о + t_в + t_обс + t_отд
  • t_о — основное (машинное) время: для точения/фрезерования t_о = L·i/(n·S), для сварки t_о = L_шв/Vсв
  • t_в — вспомогательное: установ/снятие детали, управление станком, измерения (0.3–0.5×t_о)
  • t_обс — обслуживание рабочего места (3–5% от t_о+t_в)
  • t_отд — отдых и личные нужды (4–8% от t_о+t_в)
- Для каждой операции подставляй реальные данные из чертежа: L — длина прохода, d — диаметр, Ra, квалитет
- В обосновании ВСЕГДА указывай конкретные параметры расчёта: длины, диаметры, режимы, число проходов

Нормы времени рассчитываются по методике нормирования машиностроительных предприятий:
- t_шт (штучное время, мин/дет) — основное + вспомогательное + время обслуживания рабочего места + время на отдых
- t_пз (подготовительно-заготовительное время) — возвращай ПОЛНОЕ t_пз на всю партию (мин/партию), НЕ делённое на размер партии. Фронтенд сам пересчитает t_пз на деталь.

При расчёте ОБЯЗАТЕЛЬНО учитывай:
- Материал детали и его обрабатываемость (коэффициент Kм)
- Вид обработки (токарная, фрезерная, сверлильная, шлифовальная, сварочная, слесарная и т.д.)
- Сложность геометрии, точность (квалитет, допуски), шероховатость поверхностей
- Тип и модель оборудования — подбирай режимы резания/сварки исходя из паспортных данных станка
- Количество переходов, установов, инструментальных позиций

ВЫБОР ОБОРУДОВАНИЯ — КРИТИЧЕСКИ ВАЖНЫЙ РАЗДЕЛ:
В запросе будет передана база реального оборудования завода. Рекомендуй ТОЛЬКО из этого списка.

Алгоритм выбора оптимального станка для каждой операции:
1. Определи тип операции (токарная, фрезерная, сварка и т.д.)
2. Определи РАБОЧИЙ ЦЕХ из маршрутной карты (обычно указан в шапке МК или в колонке «Цех»).
   Если цех явно не указан — считай рабочим цехом тот, что указан у оборудования первой операции.
3. Проанализируй параметры детали из чертежа и МК:
   - Точность: квалитет (IT), допуски на размеры, шероховатость Ra
   - Габариты: длина, диаметр, масса заготовки
   - Сложность геометрии: простая/сложная форма, количество поверхностей
   - Материал и его обрабатываемость
4. Выбери ОПТИМАЛЬНЫЙ станок, соблюдая ПРИОРИТЕТ ЦЕХА:
   ШАГ А — ищи подходящий станок СНАЧАЛА в рабочем цеху детали (из шага 2).
   ШАГ Б — только если нужного типа оборудования в рабочем цеху НЕТ → ищи в других цехах.
   ПРАВИЛО ПОКРАСКИ: для покрасочных операций (грунтование, окраска, лакирование, нанесение покрытий)
   все покрасочные камеры находятся в Цех №3 — ВСЕГДА выбирай оборудование только из Цех №3, даже если
   рабочий цех детали другой. В обосновании укажи: «Покрасочные камеры расположены в Цех №3».
   При этом соблюдай правила подбора:
   - Ra ≤ 1.6 или квалитет IT7 и точнее → предпочти станок с ЧПУ
   - Сложная геометрия, много переходов → станок с ЧПУ или ОЦ
   - Простая деталь, единичное производство → универсальный станок
   - Крупная деталь → проверь габариты станка (указаны в названии если есть)
   - Сварка длинных швов на большой детали → позиционер + сварочный аппарат
5. В поле "оборудование" укажи конкретное наименование из базы с указанием цеха в скобках,
   например: «Токарный станок 16К20 [Цех №1]».
   В обосновании ОБЯЗАТЕЛЬНО объясни: (а) почему выбран этот тип станка (точность, габариты),
   и (б) если станок взят из другого цеха — почему не нашлось в рабочем цеху.
6. Если в МК указана модель отсутствующая в базе — выбери аналог из базы и укажи в обосновании.

Поле "режимы" — ОБЯЗАТЕЛЬНО в каждом объекте JSON:
- Металлорежущие станки (токарные, фрезерные, сверлильные): V=... м/мин, S=... мм/об, t=... мм, n=... об/мин
- Шлифовальные: Vк=... м/с, Sпр=... м/мин, t=... мм
- Сварочные (ВСЕ виды: прихватка, прихваточные швы, сварка полуавтоматом MIG/MAG, ручная дуговая, аргонодуговая TIG, плазменная и т.д.) — ВСЕГДА требуют режимы: I=... А, U=... В, Vсв=... м/ч, d=... мм (диаметр проволоки/электрода). Прихватка — это тоже сварка, режимы обязательны!
- Прессовое: усилие=... кН, ходов/мин=...
- Термообработка: T=... °C, выдержка=... мин
- Покраска (все виды — грунтование, окраска, лакирование): давление_распыла=... бар, вязкость=... сек (ВЗ-4), расход_краски=... г/м², толщина_слоя=... мкм, число_слоёв=...
- Очистка дробеметная: скорость_дроби=... м/с, дробь=Ø...мм (тип: стальная/чугунная), время_цикла=... мин, степень_очистки=Sa... (по ISO 8501)
- Очистка пескоструйная: давление=... бар, абразив=... (тип/фракция), время_цикла=... мин, степень_очистки=Sa...
- Зачистка: инструмент=... (УШМ/напильник/щётка), P=... об/мин — или "—" если полностью вручную без оборудования
- Операции БЕЗ оборудования (комплектовочная, контрольная, маркировка, рихтовка вручную): "режимы": "—"
- Если точных данных нет — указывай типовые расчётные значения с пометкой "(расч.)"

Формат ответа — строго JSON объект:
{
  "маршрут": {
    "номер": "M-0107",
    "операции": "Комплектовочная | Прихватка | Сварка полуавтоматическая | ...",
    "источник": "типовой каталог" или "маршрутная карта",
    "обоснование": "Подробное объяснение почему выбран именно этот маршрут: тип детали, материал, характер обработки, какие признаки чертежа привели к этому выбору, какие альтернативные маршруты рассматривались и почему отклонены",
    "уверенность": 85,
    "уверенность_пояснение": "Что повышает уверенность (явные признаки) и что снижает (неоднозначности, отсутствующие данные)"
  },
  "операции": [
    {
      "деталь": "название детали из чертежа",
      "операция": "010 Токарная",
      "оборудование": "конкретное наименование станка/аппарата из базы оборудования завода",
      "t_шт_предложено": 12.5,
      "t_пз_предложено": 15.0,
      "режимы": "V=120 м/мин, S=0.3 мм/об, t=1.5 мм, n=850 об/мин",
      "обоснование": "краткое обоснование: тип обработки, режимы, основание для нормы"
    }
  ]
}

Поле "уверенность" — целое число от 0 до 100:
- 90–100: все признаки однозначны, маршрут очевиден
- 70–89: маршрут подходит хорошо, есть незначительные неоднозначности
- 50–69: несколько подходящих маршрутов, выбран наиболее вероятный
- ниже 50: данных недостаточно, выбор условный

Если маршрутная карта предоставлена (РЕЖИМ А) — в поле "источник" укажи "маршрутная карта", "номер" = "—", "уверенность" = 100, обоснование = "Маршрут взят из предоставленной маршрутной карты".

Важно:
- Только JSON, никакого текста до или после
- Числа — дробные, в минутах, с точностью до 0.1
- Поле "оборудование" присутствует ВСЕГДА — конкретный станок/аппарат из базы, или "—" для ручных операций
- Поле "режимы" присутствует ВСЕГДА — либо с параметрами, либо "—"
- Если данных недостаточно — используй типовые нормы с пометкой в обосновании
- Все строки на русском языке
"""


DRAWING_SYSTEM_PROMPT = """Ты — опытный конструктор-технолог на машиностроительном предприятии (стаж 20+ лет).
Твоя задача — провести детальный технический анализ чертежа детали и выявить:
1. ЗАМЕЧАНИЯ — ошибки, нарушения ГОСТ, отсутствующие данные, некорректные обозначения
2. ОПТИМИЗАЦИИ — предложения по улучшению технологичности, снижению стоимости изготовления

Анализируй по следующим категориям:
- допуски: квалитеты, поля допусков, предельные отклонения по ГОСТ 25347
- шероховатость: Ra/Rz, обозначения по ГОСТ 2.309, соответствие квалитету
- геометрические допуски: форма, расположение, биение по ГОСТ 2.308
- технологичность: радиусы скругления, уклоны, выходы инструмента, нетехнологичные элементы
- материал: марка стали/сплава, термообработка, покрытие, твёрдость
- оформление: основная надпись по ГОСТ 2.104, технические требования, масштаб, виды, разрезы по ГОСТ 2.305

Для каждого замечания/оптимизации укажи ПРИБЛИЗИТЕЛЬНОЕ расположение на чертеже в процентах (0–100 от левого верхнего угла листа).
Оценивай координаты по типовым зонам:
- основная надпись — правый нижний угол (x≈65–100%, y≈80–100%)
- технические требования — верхний левый угол (x≈2–35%, y≈3–20%)
- главный вид — центр листа (x≈30–70%, y≈25–75%)
- разрез/вид слева — правая часть (x≈65–95%, y≈25–65%)
- вид сверху — нижняя центральная часть (x≈25–65%, y≈60–85%)

ВАЖНО: возвращай СТРОГО JSON-объект, без пояснений вне JSON.

Формат ответа:
{
  "summary": "Краткое общее резюме по чертежу (2–3 предложения): общее впечатление, критичность найденных замечаний",
  "remarks": [
    {
      "id": 1,
      "type": "замечание",
      "category": "допуски",
      "priority": "высокий",
      "title": "Краткое название (до 60 символов)",
      "description": "Подробное описание проблемы с указанием конкретного элемента чертежа и ссылкой на ГОСТ если применимо",
      "suggestion": "Конкретная рекомендация: что именно добавить/изменить/исправить",
      "x": 35.0,
      "y": 40.0,
      "w": 20.0,
      "h": 10.0
    }
  ]
}

Правила:
- x, y, w, h — проценты от размера листа (0–100); x,y — левый верхний угол области, w,h — ширина и высота
- type: "замечание" (ошибки, несоответствия ГОСТ) или "оптимизация" (улучшения, предложения)
- priority: "высокий" (критично для функционирования/изготовления), "средний" (желательно исправить), "низкий" (незначительное замечание)
- category: "допуски", "шероховатость", "геометрические допуски", "технологичность", "материал", "оформление", "другое"
- Минимум 3 замечания/оптимизации, максимум 15
- Все строки на русском языке
"""


def analyze_drawing_with_claude(chertezh):
    """Анализирует чертёж через Anthropic API, возвращает замечания и оптимизации."""
    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        raise ValueError("ANTHROPIC_API_KEY не задан в .env")

    chertezh_b64 = pdf_to_base64(chertezh)

    content = [
        {"type": "text", "text": "Проведи детальный технический анализ этого чертежа детали:"},
        {
            "type": "document",
            "source": {"type": "base64", "media_type": "application/pdf", "data": chertezh_b64},
        },
        {"type": "text", "text": (
            "Выяви все замечания (ошибки, нарушения ГОСТ) и оптимизации (предложения по улучшению технологичности).\n"
            "Для каждого пункта укажи приблизительные координаты на чертеже (x,y,w,h в процентах от листа).\n"
            "Верни строго JSON-объект по указанному формату."
        )},
    ]

    client = Anthropic(api_key=api_key)
    message = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=4096,
        system=DRAWING_SYSTEM_PROMPT,
        messages=[{"role": "user", "content": content}],
    )

    response_text = message.content[0].text.strip()
    if response_text.startswith("```"):
        lines = response_text.split("\n")
        response_text = "\n".join(lines[1:-1])

    return json.loads(response_text)


def analyze_drawing_with_claude_code(chertezh_file):
    """Анализирует чертёж через Claude Code CLI."""
    tmp1 = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f1:
            chertezh_file.save(f1.name)
            tmp1 = f1.name

        prompt = (
            f"Прочитай PDF-файл чертежа детали: {tmp1}\n\n"
            f"Проведи детальный технический анализ:\n"
            f"1. Найди замечания (ошибки, нарушения ГОСТ, отсутствующие данные)\n"
            f"2. Найди оптимизации (предложения по улучшению технологичности)\n"
            f"Для каждого пункта укажи приблизительные координаты на листе (x,y,w,h в процентах 0-100).\n"
            f"Верни ТОЛЬКО JSON-объект по формату из системного промпта."
        )

        env = os.environ.copy()
        env.pop("CLAUDECODE", None)

        result = subprocess.run(
            [
                CLAUDE_BIN,
                "-p", prompt,
                "--allowedTools", "Read",
                "--output-format", "json",
                "--system-prompt", DRAWING_SYSTEM_PROMPT,
                "--model", "sonnet",
                "--no-session-persistence",
            ],
            capture_output=True,
            text=True,
            timeout=300,
            env=env,
        )

        if result.returncode != 0:
            raise RuntimeError(f"Claude Code вернул ошибку: {result.stderr[:500]}")

        data = json.loads(result.stdout)
        if data.get("is_error"):
            raise RuntimeError(f"Claude Code: {data.get('result', 'неизвестная ошибка')}")

        response_text = data["result"].strip()
        if response_text.startswith("```"):
            lines = response_text.split("\n")
            response_text = "\n".join(lines[1:-1])

        return json.loads(response_text)

    finally:
        if tmp1 and os.path.exists(tmp1):
            os.unlink(tmp1)


def get_db():
    db = getattr(g, "_database", None)
    if db is None:
        db = g._database = sqlite3.connect(DB_PATH)
        db.row_factory = sqlite3.Row
    return db


@app.teardown_appcontext
def close_db(exception):
    db = getattr(g, "_database", None)
    if db is not None:
        db.close()


def init_db():
    with app.app_context():
        db = sqlite3.connect(DB_PATH)
        db.execute("""
            CREATE TABLE IF NOT EXISTS нормы (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                деталь TEXT NOT NULL,
                операция TEXT NOT NULL,
                t_шт_предложено REAL,
                t_шт_подтверждено REAL,
                t_пз_предложено REAL,
                t_пз_подтверждено REAL,
                режимы TEXT,
                обоснование TEXT,
                дата TEXT,
                подтверждено INTEGER DEFAULT 0
            )
        """)
        # Миграция: добавляем колонку режимы если её нет (для существующих БД)
        try:
            db.execute("ALTER TABLE нормы ADD COLUMN режимы TEXT")
        except Exception:
            pass  # Колонка уже существует
        try:
            db.execute("ALTER TABLE нормы ADD COLUMN изделие TEXT")
        except Exception:
            pass
        try:
            db.execute("ALTER TABLE нормы ADD COLUMN оборудование TEXT")
        except Exception:
            pass
        db.commit()
        db.close()


def pdf_to_base64(file_storage):
    return base64.standard_b64encode(file_storage.read()).decode("utf-8")


@app.route("/")
def index():
    return render_template("index.html")


STUB_OPERATIONS = [
    {
        "деталь": "Вал ступенчатый (тестовые данные)",
        "операция": "010 Токарная",
        "t_шт_предложено": 12.5,
        "t_пз_предложено": 15.0,
        "режимы": "V=120 м/мин, S=0.3 мм/об, t=1.5 мм, n=850 об/мин",
        "обоснование": "Черновое и чистовое точение Ø45h6, L=320мм, сталь 45, Ra 1.6"
    },
    {
        "деталь": "Вал ступенчатый (тестовые данные)",
        "операция": "020 Фрезерная",
        "t_шт_предложено": 8.3,
        "t_пз_предложено": 12.0,
        "режимы": "V=80 м/мин, Sz=0.05 мм/зуб, t=5.5 мм, n=1820 об/мин",
        "обоснование": "Фрезерование шпоночного паза 14×5.5мм, концевая фреза Ø14"
    },
    {
        "деталь": "Вал ступенчатый (тестовые данные)",
        "операция": "030 Сверлильная",
        "t_шт_предложено": 4.2,
        "t_пз_предложено": 8.0,
        "режимы": "V=25 м/мин, S=0.12 мм/об, n=1600 об/мин",
        "обоснование": "Сверление центрового отверстия Ø5мм, нарезание резьбы М8"
    },
    {
        "деталь": "Вал ступенчатый (тестовые данные)",
        "операция": "040 Шлифовальная",
        "t_шт_предложено": 18.0,
        "t_пз_предложено": 20.0,
        "режимы": "Vк=35 м/с, Sпр=1.5 м/мин, t=0.02 мм, n=200 об/мин",
        "обоснование": "Круглое шлифование шеек Ø45h6, Ra 0.8, квалитет h6"
    },
]


def analyze_with_claude_code(chertezh_file, marshrutnaya_file=None, batch_size=1, tipovoy_file=None):
    """Обрабатывает PDF через Claude Code CLI (claude -p)."""
    fname = getattr(chertezh_file, 'filename', 'чертёж') or 'чертёж'
    alog(fname, f"Начало анализа (Claude Code CLI): {fname}")
    tmp1 = tmp2 = tmp3 = None
    try:
        # Сохраняем загруженные файлы во временные PDF
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f1:
            chertezh_file.save(f1.name)
            tmp1 = f1.name
        if marshrutnaya_file:
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f2:
                marshrutnaya_file.save(f2.name)
                tmp2 = f2.name
        if tipovoy_file:
            with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as f3:
                tipovoy_file.save(f3.name)
                tmp3 = f3.name

        alog(fname, "Загрузка базы оборудования завода...")
        equipment_text = load_equipment_text()
        equipment_section = f"\n{equipment_text}\n" if equipment_text else ""
        alog(fname, f"База оборудования: {len(equipment_text)} символов" if equipment_text else "База оборудования пуста")

        if tmp2:
            # РЕЖИМ А: чертёж + маршрутная карта
            num = 2
            chertezh_part = f"1. Чертёж детали: {tmp1}\n"
            mk_part = f"2. Маршрутная карта: {tmp2}\n"
            tipovoy_part = (
                f"3. Типовой техпроцесс: {tmp3}\n" if tmp3 else ""
            )
            mode_instruction = "Нормируй операции из маршрутной карты, используя чертёж для расчёта параметров."
            typical_section = ""
        else:
            # РЕЖИМ Б: только чертёж → выбрать типовой маршрут и нормировать
            chertezh_part = f"1. Чертёж детали: {tmp1}\n"
            mk_part = ""
            tipovoy_part = ""
            mode_instruction = (
                "РЕЖИМ Б: Маршрутная карта НЕ предоставлена.\n"
                "1. Проанализируй чертёж: тип детали, материал, геометрия, точность.\n"
                "2. Выбери НАИБОЛЕЕ ПОДХОДЯЩИЙ типовой маршрут из каталога ниже.\n"
                "3. В поле 'обоснование' первой операции укажи: «Выбран маршрут M-XXXX: [операции]».\n"
                "4. Пронормируй каждую операцию выбранного маршрута.\n"
                "Нумеруй операции: 010, 020, 030..."
            )
            typical_routes = load_typical_routes()
            typical_section = f"\n{typical_routes}\n" if typical_routes else ""

        prompt = (
            f"Прочитай PDF файлы:\n"
            f"{chertezh_part}"
            f"{mk_part}"
            f"{tipovoy_part}\n"
            f"{mode_instruction}\n\n"
            f"Размер партии: {batch_size} деталей.\n"
            f"{equipment_section}"
            f"{typical_section}\n"
            f"Рассчитай нормы времени для всех операций.\n\n"
            f"ОБЯЗАТЕЛЬНО: в поле 'режимы' каждого объекта JSON укажи конкретные параметры оборудования.\n"
            f"Для сварочных операций (прихватка, сварка, наплавка): I=...А, U=...В, Vсв=...м/ч, d=...мм\n"
            f"Для металлорежущих станков: V=...м/мин, S=...мм/об, t=...мм, n=...об/мин\n"
            f"Для ручных операций (комплектовочная, контрольная): режимы='—'\n\n"
            f"Верни ТОЛЬКО JSON массив, без пояснений."
        )

        alog(fname, "Отправка запроса в Claude Code CLI...")
        env = os.environ.copy()
        env.pop("CLAUDECODE", None)  # разрешаем вложенный запуск

        result = subprocess.run(
            [
                CLAUDE_BIN,
                "-p", prompt,
                "--allowedTools", "Read",
                "--output-format", "json",
                "--system-prompt", SYSTEM_PROMPT,
                "--model", "sonnet",
                "--no-session-persistence",
            ],
            capture_output=True,
            text=True,
            timeout=300,
            env=env,
        )

        if result.returncode != 0:
            alog(fname, f"Ошибка Claude Code CLI (код {result.returncode}): {result.stderr[:200]}", "error")
            raise RuntimeError(f"Claude Code вернул ошибку: {result.stderr[:500]}")

        alog(fname, "Ответ получен, разбор JSON...")
        # Формат --output-format json: {"result": "...", "is_error": false, ...}
        data = json.loads(result.stdout)
        if data.get("is_error"):
            alog(fname, f"Claude Code вернул ошибку: {data.get('result', '?')[:200]}", "error")
            raise RuntimeError(f"Claude Code: {data.get('result', 'неизвестная ошибка')}")

        response_text = data["result"].strip()
        if response_text.startswith("```"):
            lines = response_text.split("\n")
            response_text = "\n".join(lines[1:-1])

        parsed = json.loads(response_text)
        # Поддержка нового формата {маршрут, операции} и старого формата (массив)
        if isinstance(parsed, dict) and "операции" in parsed:
            маршрут = parsed.get("маршрут", {})
            операции = parsed["операции"]
        else:
            маршрут = {}
            операции = parsed
        alog(fname, f"Результат: {len(операции)} операций")
        for op in операции:
            alog(fname, f"  {op.get('операция','?')} → t_шт={op.get('t_шт_предложено','?')}, оборуд={op.get('оборудование','?')}")
            print(f"[DEBUG claude-code] {op.get('операция','?')[:30]} -> режимы={repr(op.get('режимы','ОТСУТСТВУЕТ'))}", flush=True)
        alog(fname, "Анализ завершён", "success")
        return {"маршрут": маршрут, "операции": операции}

    finally:
        if tmp1 and os.path.exists(tmp1):
            os.unlink(tmp1)
        if tmp2 and os.path.exists(tmp2):
            os.unlink(tmp2)
        if tmp3 and os.path.exists(tmp3):
            os.unlink(tmp3)


def analyze_with_claude(chertezh, marshrutnaya=None, batch_size=1, tipovoy=None):
    fname = getattr(chertezh, 'filename', 'чертёж') or 'чертёж'
    alog(fname, f"Начало анализа (Anthropic API): {fname}")
    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        alog(fname, "ANTHROPIC_API_KEY не задан!", "error")
        raise ValueError("ANTHROPIC_API_KEY не задан в .env")

    alog(fname, "Кодирование PDF в base64...")
    chertezh_b64 = pdf_to_base64(chertezh)
    content = [
        {"type": "text", "text": "Файл 1 — чертёж детали:"},
        {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": chertezh_b64}},
    ]

    if marshrutnaya:
        marshrutnaya_b64 = pdf_to_base64(marshrutnaya)
        content.append({"type": "text", "text": "Файл 2 — маршрутная карта:"})
        content.append({
            "type": "document",
            "source": {"type": "base64", "media_type": "application/pdf", "data": marshrutnaya_b64},
        })

    if tipovoy:
        tipovoy_b64 = pdf_to_base64(tipovoy)
        content.append({"type": "text", "text": "Типовой техпроцесс (эталонный перечень операций):"})
        content.append({
            "type": "document",
            "source": {"type": "base64", "media_type": "application/pdf", "data": tipovoy_b64},
        })

    alog(fname, "Загрузка базы оборудования завода...")
    equipment_text = load_equipment_text()
    equipment_section = f"\n{equipment_text}\n\n" if equipment_text else ""
    alog(fname, f"База оборудования: {len(equipment_text)} символов" if equipment_text else "База оборудования пуста")

    if marshrutnaya:
        mode_instruction = "Нормируй операции из маршрутной карты, используя чертёж для расчёта параметров."
        typical_section = ""
    else:
        mode_instruction = (
            "РЕЖИМ Б: Маршрутная карта НЕ предоставлена.\n"
            "1. Проанализируй чертёж: тип детали, материал, геометрия, точность.\n"
            "2. Выбери НАИБОЛЕЕ ПОДХОДЯЩИЙ типовой маршрут из каталога ниже.\n"
            "3. В поле 'обоснование' первой операции укажи: «Выбран маршрут M-XXXX: [операции]».\n"
            "4. Пронормируй каждую операцию выбранного маршрута.\n"
            "Нумеруй операции: 010, 020, 030..."
        )
        typical_routes = load_typical_routes()
        typical_section = f"\n{typical_routes}\n\n" if typical_routes else ""

    content.append({"type": "text", "text": (
        f"{mode_instruction}\n\n"
        f"Размер партии: {batch_size} деталей.\n"
        f"{equipment_section}"
        f"{typical_section}"
        f"Рассчитай нормы времени для всех операций.\n\n"
        f"ОБЯЗАТЕЛЬНО: в поле 'режимы' каждого объекта JSON укажи конкретные параметры оборудования.\n"
        f"Для сварочных операций (прихватка, сварка, наплавка): I=...А, U=...В, Vсв=...м/ч, d=...мм\n"
        f"Для металлорежущих станков: V=...м/мин, S=...мм/об, t=...мм, n=...об/мин\n"
        f"Для ручных операций (комплектовочная, контрольная): режимы='—'\n\n"
        f"Верни только JSON."
    )})

    alog(fname, "Отправка запроса в Anthropic API (claude-sonnet-4)...")
    client = Anthropic(api_key=api_key)
    message = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=4096,
        system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": content}],
    )

    alog(fname, f"Ответ получен: {message.usage.input_tokens} вх. токенов, {message.usage.output_tokens} вых. токенов")

    response_text = message.content[0].text.strip()
    if response_text.startswith("```"):
        lines = response_text.split("\n")
        response_text = "\n".join(lines[1:-1])

    alog(fname, "Разбор JSON ответа...")
    parsed = json.loads(response_text)
    if isinstance(parsed, dict) and "операции" in parsed:
        маршрут = parsed.get("маршрут", {})
        result_data = parsed["операции"]
    else:
        маршрут = {}
        result_data = parsed
    alog(fname, f"Результат: {len(result_data)} операций")
    for op in result_data:
        alog(fname, f"  {op.get('операция','?')} → t_шт={op.get('t_шт_предложено','?')}, оборуд={op.get('оборудование','?')}")
        print(f"[DEBUG api] {op.get('операция','?')[:30]} -> режимы={repr(op.get('режимы','ОТСУТСТВУЕТ'))}", flush=True)
    alog(fname, "Анализ завершён", "success")
    return {"маршрут": маршрут, "операции": result_data}


@app.route("/api/analyze", methods=["POST"])
def analyze():
    chertezh = request.files.get("chertezh")
    if not chertezh or chertezh.filename == "":
        return jsonify({"error": "Необходимо загрузить чертёж детали"}), 400

    marshrutnaya = request.files.get("marshrutnaya")
    if marshrutnaya and marshrutnaya.filename == "":
        marshrutnaya = None

    batch_size = request.form.get("batch_size", 1)
    try:
        batch_size = int(batch_size)
        if batch_size < 1:
            batch_size = 1
    except (ValueError, TypeError):
        batch_size = 1

    tipovoy = request.files.get("tipovoy")

    use_stub = os.getenv("USE_STUB", "false").lower() == "true"
    use_claude_code = os.getenv("USE_CLAUDE_CODE", "false").lower() == "true"

    try:
        if use_stub:
            alog(chertezh.filename, "Режим STUB — возвращаем тестовые данные", "warn")
            result = {"маршрут": {}, "операции": STUB_OPERATIONS}
        elif use_claude_code:
            result = analyze_with_claude_code(chertezh, marshrutnaya, batch_size, tipovoy)
        else:
            result = analyze_with_claude(chertezh, marshrutnaya, batch_size, tipovoy)

        return jsonify({
            "operations": result["операции"],
            "маршрут": result.get("маршрут", {}),
            "stub": use_stub
        })

    except json.JSONDecodeError as e:
        return jsonify({"error": f"Не удалось разобрать JSON от API: {str(e)}", "raw": response_text}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/analyze_drawing", methods=["POST"])
def analyze_drawing_route():
    chertezh = request.files.get("chertezh")
    if not chertezh or chertezh.filename == "":
        return jsonify({"error": "Необходимо загрузить чертёж детали"}), 400

    use_stub = os.getenv("USE_STUB", "false").lower() == "true"
    use_claude_code = os.getenv("USE_CLAUDE_CODE", "false").lower() == "true"

    try:
        if use_stub:
            result = {
                "summary": "Тестовый анализ чертежа. Обнаружено 3 замечания и 2 оптимизации.",
                "remarks": [
                    {"id": 1, "type": "замечание", "category": "допуски", "priority": "высокий",
                     "title": "Не указан допуск на размер Ø45", "x": 35.0, "y": 40.0, "w": 18.0, "h": 8.0,
                     "description": "Для поверхности Ø45 не указан квалитет и поле допуска. Нарушение ГОСТ 2.307.",
                     "suggestion": "Добавить обозначение Ø45h6 или Ø45±0.012 в зависимости от требуемой посадки."},
                    {"id": 2, "type": "замечание", "category": "шероховатость", "priority": "средний",
                     "title": "Шероховатость не соответствует квалитету", "x": 55.0, "y": 30.0, "w": 15.0, "h": 10.0,
                     "description": "На поверхности IT6 указана шероховатость Ra 3.2, что не соответствует требованиям.",
                     "suggestion": "Для квалитета IT6 установить Ra ≤ 1.6 мкм."},
                    {"id": 3, "type": "оптимизация", "category": "технологичность", "priority": "средний",
                     "title": "Нетехнологичный радиус R0.5", "x": 45.0, "y": 60.0, "w": 12.0, "h": 8.0,
                     "description": "Радиус скругления R0.5 сложен в изготовлении и требует специального инструмента.",
                     "suggestion": "Увеличить до R1.0 — стандартный инструмент, снижение стоимости ~15%."},
                ],
            }
        elif use_claude_code:
            result = analyze_drawing_with_claude_code(chertezh)
        else:
            result = analyze_drawing_with_claude(chertezh)

        return jsonify(result)

    except json.JSONDecodeError as e:
        return jsonify({"error": f"Не удалось разобрать JSON от API: {str(e)}"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/confirm", methods=["POST"])
def confirm():
    data = request.get_json()
    if not data or "operations" not in data:
        return jsonify({"error": "Нет данных для сохранения"}), 400

    operations = data["operations"]
    изделие = data.get("изделие", "")
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    db = get_db()
    saved_ids = []

    try:
        for op in operations:
            cursor = db.execute(
                """
                INSERT INTO нормы
                    (деталь, операция, оборудование, t_шт_предложено, t_шт_подтверждено,
                     t_пз_предложено, t_пз_подтверждено, режимы, обоснование, дата, подтверждено, изделие)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?)
                """,
                (
                    op.get("деталь", ""),
                    op.get("операция", ""),
                    op.get("оборудование", ""),
                    op.get("t_шт_предложено"),
                    op.get("t_шт_подтверждено", op.get("t_шт_предложено")),
                    op.get("t_пз_предложено"),
                    op.get("t_пз_подтверждено", op.get("t_пз_предложено")),
                    op.get("режимы", ""),
                    op.get("обоснование", ""),
                    now,
                    изделие,
                ),
            )
            saved_ids.append(cursor.lastrowid)
        db.commit()
        return jsonify({"saved": len(saved_ids), "ids": saved_ids})
    except Exception as e:
        db.rollback()
        return jsonify({"error": str(e)}), 500


@app.route("/api/logs", methods=["GET"])
def get_logs():
    """Возвращает лог анализа. ?since=N — только записи начиная с индекса N."""
    since = request.args.get("since", 0, type=int)
    with _analysis_logs_lock:
        entries = _analysis_logs[since:]
        total = len(_analysis_logs)
    return jsonify({"logs": entries, "total": total})


@app.route("/api/logs/clear", methods=["POST"])
def clear_logs():
    """Очищает лог."""
    with _analysis_logs_lock:
        _analysis_logs.clear()
    return jsonify({"ok": True})


@app.route("/api/history", methods=["GET"])
def history():
    db = get_db()
    rows = db.execute(
        "SELECT * FROM нормы ORDER BY дата DESC LIMIT 200"
    ).fetchall()
    return jsonify([dict(r) for r in rows])


# ─── Анализ PDF с диска (для каталога изделий) ─────────────────────────────────

PRODUCTS_CACHE_FILE = "_norming_results.json"


def _load_products_cache(variant_path):
    """Загружает кеш результатов анализа из папки варианта."""
    cache_path = os.path.join(variant_path, PRODUCTS_CACHE_FILE)
    if os.path.isfile(cache_path):
        try:
            with open(cache_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, OSError):
            pass
    return {}


def _save_products_cache(variant_path, cache):
    """Сохраняет кеш результатов анализа в папку варианта."""
    cache_path = os.path.join(variant_path, PRODUCTS_CACHE_FILE)
    with open(cache_path, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)


def analyze_pdf_from_disk(pdf_path, batch_size=1):
    """Анализирует PDF-файл с диска — та же логика, что analyze_with_claude, но файл уже на диске."""
    fname = os.path.basename(pdf_path)
    use_stub = os.getenv("USE_STUB", "false").lower() == "true"
    use_claude_code = os.getenv("USE_CLAUDE_CODE", "false").lower() == "true"

    alog(fname, f"Начало анализа файла: {fname}")
    alog(fname, f"Размер файла: {os.path.getsize(pdf_path)} байт")

    if use_stub:
        alog(fname, "Режим STUB — возвращаем тестовые данные", "warn")
        return {"маршрут": {}, "операции": STUB_OPERATIONS}

    method = "Claude Code CLI" if use_claude_code else "Anthropic REST API"
    alog(fname, f"Метод анализа: {method}")

    if use_claude_code:
        return _analyze_disk_claude_code(pdf_path, batch_size)
    else:
        return _analyze_disk_claude_api(pdf_path, batch_size)


def _run_claude_cli(prompt, system_prompt=None, allowed_tools=None, timeout=120):
    """Вспомогательная функция запуска Claude Code CLI."""
    env = os.environ.copy()
    env.pop("CLAUDECODE", None)
    cmd = [CLAUDE_BIN, "-p", prompt, "--output-format", "json", "--model", "sonnet", "--no-session-persistence"]
    if system_prompt:
        cmd += ["--system-prompt", system_prompt]
    if allowed_tools:
        cmd += ["--allowedTools", allowed_tools]
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout, env=env)
    if result.returncode != 0:
        raise RuntimeError(f"Claude Code ошибка (код {result.returncode}): {result.stderr[:300]}")
    # Парсим stdout — может быть JSON от CLI или прямой текст
    stdout = result.stdout.strip()
    if not stdout:
        raise RuntimeError("Claude Code вернул пустой ответ")
    try:
        data = json.loads(stdout)
        if data.get("is_error"):
            raise RuntimeError(f"Claude Code: {data.get('result', 'ошибка')[:300]}")
        text = data.get("result", "").strip()
    except json.JSONDecodeError:
        text = stdout  # fallback — ответ не в JSON-обёртке
    if text.startswith("```"):
        lines = text.split("\n")
        text = "\n".join(lines[1:-1])
    return text


def _load_equipment_by_ops():
    """Загружает оборудование как словарь {тип_операции: текст}."""
    if not _OPENPYXL_AVAILABLE or not os.path.exists(EQUIPMENT_DB_PATH):
        return {}
    wb = openpyxl.load_workbook(EQUIPMENT_DB_PATH, read_only=True, data_only=True)
    ws = wb['Лист1']
    by_op = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[1]:
            continue
        name = str(row[1]).replace('\xa0', ' ').strip()
        цех = str(row[2]).strip() if row[2] else '—'
        ops = [str(x).strip() for x in [row[4], row[5], row[6]] if x]
        for op in ops:
            by_op.setdefault(op, []).append(f"  • {name[:80]} [{цех}]")
    wb.close()
    return by_op


def _filter_equipment(operation_names, eq_by_ops):
    """Фильтрует оборудование только по нужным операциям."""
    lines = ["ОБОРУДОВАНИЕ ЗАВОДА (только для указанных операций):"]
    matched = set()
    for op_need in operation_names:
        op_lower = op_need.lower().strip()
        for op_key, items in eq_by_ops.items():
            if op_lower in op_key.lower() or op_key.lower() in op_lower:
                if op_key not in matched:
                    matched.add(op_key)
                    lines.append(f"\n{op_key}:")
                    seen = set()
                    for item in items:
                        if item not in seen:
                            seen.add(item)
                            lines.append(item)
    if len(matched) == 0:
        return load_equipment_text()  # fallback — вся база
    return "\n".join(lines)


def _extract_json_value(text, opener='['):
    """Находит первый корректный JSON-массив или объект с учётом вложенных скобок."""
    closer = ']' if opener == '[' else '}'
    start = text.find(opener)
    if start == -1:
        return None
    depth = 0
    in_str = False
    escape = False
    for i, ch in enumerate(text[start:], start):
        if escape:
            escape = False
            continue
        if ch == '\\' and in_str:
            escape = True
            continue
        if ch == '"':
            in_str = not in_str
            continue
        if in_str:
            continue
        if ch == opener:
            depth += 1
        elif ch == closer:
            depth -= 1
            if depth == 0:
                return text[start:i + 1]
    return None


def _extract_json_array(text):
    return _extract_json_value(text, '[')


def _is_assembly_drawing(filename):
    """Проверяет, является ли файл сборочным чертежом (СБ)."""
    import re as _re
    name = os.path.basename(filename)
    return bool(_re.search(r'\bСБ\b', name, _re.IGNORECASE) or _re.search(r'сборочн', name, _re.IGNORECASE))


def _analyze_disk_claude_code(pdf_path, batch_size=1):
    """Двухшаговый анализ PDF через Claude Code CLI."""
    fname = os.path.basename(pdf_path)
    is_assembly = _is_assembly_drawing(fname)

    # ═══ ШАГ 1: Быстрый запрос — определить деталь и операции ═══
    alog(fname, "Шаг 1: определение детали и списка операций...")
    if is_assembly:
        alog(fname, "Сборочный чертёж (СБ) — запрашиваем только операции уровня сборки")
        step1_prompt = (
            f"Прочитай PDF файл: {pdf_path}\n\n"
            f"Это СБОРОЧНЫЙ чертёж (СБ). Определи:\n"
            f"1. Название изделия/сборочной единицы\n"
            f"2. Материал или основной конструкционный материал (если указан)\n"
            f"3. Список операций СБОРОЧНОГО уровня — то есть операции, выполняемые при сборке готового изделия:\n"
            f"   сварка, слесарная сборка, зачистка сварных швов, покрасочная, контроль, испытания.\n"
            f"   НЕ включай операции изготовления отдельных деталей (резка, токарная, фрезерная, сверлильная и т.п.) — "
            f"они относятся к деталям, а не к сборке.\n\n"
            f"Верни ТОЛЬКО JSON:\n"
            f'{{"деталь": "название изделия", "материал": "марка", "операции": ["Сварочная", "Слесарная", "Зачистка", "Покрасочная", "Контроль"]}}'
        )
    else:
        step1_prompt = (
            f"Прочитай PDF файл: {pdf_path}\n\n"
            f"Определи:\n"
            f"1. Название детали\n"
            f"2. Материал\n"
            f"3. Список необходимых операций обработки (по типу детали и чертежу)\n\n"
            f"Верни ТОЛЬКО JSON:\n"
            f'{{"деталь": "название", "материал": "марка", "операции": ["Газо-плазменная резка", "Зачистка", "Токарная", ...]}}'
        )

    try:
        step1_text = _run_claude_cli(step1_prompt, allowed_tools="Read", timeout=90)
        # Извлекаем первый JSON-объект из ответа с учётом вложенных скобок
        json_obj = _extract_json_value(step1_text, '{')
        if json_obj:
            step1 = json.loads(json_obj)
        else:
            step1 = json.loads(step1_text)
        op_list = step1.get("операции", [])
        part_name = step1.get("деталь", fname)
        material = step1.get("материал", "")
        alog(fname, f"Шаг 1 готов: {part_name}, материал: {material}, операции: {op_list}")
    except Exception as e:
        alog(fname, f"Шаг 1 ошибка: {e}, fallback на полный промпт", "warn")
        op_list = []
        part_name = fname
        material = ""

    # ═══ ШАГ 2: Скрипт — фильтрация оборудования ═══
    alog(fname, "Шаг 2: фильтрация оборудования...")
    eq_by_ops = _load_equipment_by_ops()

    if op_list:
        filtered_eq = _filter_equipment(op_list, eq_by_ops)
    else:
        filtered_eq = load_equipment_text()  # fallback
    alog(fname, f"Оборудование после фильтрации: {len(filtered_eq)} символов (было {len(load_equipment_text())})")

    # ═══ ШАГ 3: Расчёт — компактный промпт (без каталога маршрутов — операции уже определены) ═══
    alog(fname, "Шаг 3: расчёт норм времени...")
    step3_prompt = (
        f"Прочитай PDF файл: {pdf_path}\n\n"
        f"Ты — нормировщик машиностроительного завода. Рассчитай нормы времени.\n"
        f"Деталь: {part_name}, материал: {material}\n"
        f"Операции: {', '.join(op_list) if op_list else 'определи из чертежа'}\n"
        f"Партия: {batch_size} шт.\n\n"
        f"{filtered_eq}\n\n"
        f"Для каждой операции верни JSON объект с полями:\n"
        f"деталь, операция (010 Название), оборудование (из списка выше [Цех]),\n"
        f"t_шт_предложено (мин/дет), t_пз_предложено (мин/партию),\n"
        f"режимы (сварка: I/U/Vсв/d, резка: V/S/t/n, ручные: —),\n"
        f"обоснование.\n\n"
        f"Верни ТОЛЬКО JSON массив, без пояснений."
    )

    alog(fname, f"Промпт шага 3: {len(step3_prompt)} символов")
    step3_text = _run_claude_cli(step3_prompt, allowed_tools="Read", timeout=300)
    alog(fname, f"Шаг 3 ответ получен ({len(step3_text)} символов)")

    # Надёжный парсинг — ищем первый корректный JSON массив (с учётом вложенности)
    json_arr = _extract_json_array(step3_text)
    if json_arr:
        parsed = json.loads(json_arr)
    else:
        parsed = json.loads(step3_text)
    if isinstance(parsed, dict) and "операции" in parsed:
        ops = parsed["операции"]
        route = parsed.get("маршрут", {})
    else:
        ops = parsed
        route = {}

    alog(fname, f"Результат: {len(ops)} операций")
    for op in ops:
        alog(fname, f"  {op.get('операция','?')} → t_шт={op.get('t_шт_предложено','?')}, оборуд={op.get('оборудование','?')}")
    alog(fname, "Анализ завершён", "success")
    return {"маршрут": route, "операции": ops}


def _analyze_disk_claude_api(pdf_path, batch_size=1):
    """Анализ PDF с диска через Anthropic REST API."""
    fname = os.path.basename(pdf_path)
    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        alog(fname, "ANTHROPIC_API_KEY не задан!", "error")
        raise ValueError("ANTHROPIC_API_KEY не задан в .env")

    alog(fname, "Чтение PDF и кодирование в base64...")
    with open(pdf_path, "rb") as f:
        pdf_b64 = base64.standard_b64encode(f.read()).decode("utf-8")
    alog(fname, f"PDF закодирован: {len(pdf_b64)} символов base64")

    alog(fname, "Загрузка базы оборудования завода...")
    equipment_text = load_equipment_text()
    equipment_section = f"\n{equipment_text}\n\n" if equipment_text else ""
    alog(fname, f"База оборудования: {len(equipment_text)} символов" if equipment_text else "База оборудования пуста")

    alog(fname, "Загрузка каталога типовых маршрутов...")
    typical_routes = load_typical_routes()
    typical_section = f"\n{typical_routes}\n\n" if typical_routes else ""
    alog(fname, f"Каталог маршрутов: {len(typical_routes)} символов" if typical_routes else "Каталог маршрутов пуст")

    mode_instruction = (
        "РЕЖИМ Б: Маршрутная карта НЕ предоставлена.\n"
        "1. Проанализируй чертёж/документ: тип детали, материал, геометрия, точность.\n"
        "2. Выбери НАИБОЛЕЕ ПОДХОДЯЩИЙ типовой маршрут из каталога ниже.\n"
        "3. В поле 'обоснование' первой операции укажи: «Выбран маршрут M-XXXX: [операции]».\n"
        "4. Пронормируй каждую операцию выбранного маршрута.\n"
        "Нумеруй операции: 010, 020, 030..."
    )

    content = [
        {"type": "text", "text": "Документ (чертёж или маршрутная карта):"},
        {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": pdf_b64}},
        {"type": "text", "text": (
            f"{mode_instruction}\n\n"
            f"Размер партии: {batch_size} деталей.\n"
            f"{equipment_section}"
            f"{typical_section}"
            f"Рассчитай нормы времени для всех операций.\n\n"
            f"ОБЯЗАТЕЛЬНО: в поле 'режимы' каждого объекта JSON укажи конкретные параметры оборудования.\n"
            f"Для сварочных операций (прихватка, сварка, наплавка): I=...А, U=...В, Vсв=...м/ч, d=...мм\n"
            f"Для металлорежущих станков: V=...м/мин, S=...мм/об, t=...мм, n=...об/мин\n"
            f"Для ручных операций (комплектовочная, контрольная): режимы='—'\n\n"
            f"Верни только JSON."
        )},
    ]

    alog(fname, "Отправка запроса в Anthropic API (claude-sonnet-4)...")
    client = Anthropic(api_key=api_key)
    message = client.messages.create(
        model="claude-sonnet-4-20250514",
        max_tokens=4096,
        system=SYSTEM_PROMPT,
        messages=[{"role": "user", "content": content}],
    )
    alog(fname, f"Ответ получен: {message.usage.input_tokens} вх. токенов, {message.usage.output_tokens} вых. токенов")

    response_text = message.content[0].text.strip()
    if response_text.startswith("```"):
        lines = response_text.split("\n")
        response_text = "\n".join(lines[1:-1])

    alog(fname, "Разбор JSON ответа...")
    parsed = json.loads(response_text)
    if isinstance(parsed, dict) and "операции" in parsed:
        ops = parsed["операции"]
        route = parsed.get("маршрут", {})
    else:
        ops = parsed
        route = {}

    alog(fname, f"Результат: {len(ops)} операций")
    for op in ops:
        alog(fname, f"  {op.get('операция','?')} → t_шт={op.get('t_шт_предложено','?')}, оборуд={op.get('оборудование','?')}")
    alog(fname, "Анализ завершён", "success")
    return {"маршрут": route, "операции": ops}


# ─── API каталога изделий ──────────────────────────────────────────────────────

@app.route("/api/products/tree", methods=["GET"])
def products_tree():
    """Дерево типов изделий и вариантов для сайдбара."""
    base = os.path.realpath(PRODUCTS_BASE_PATH)
    if not os.path.isdir(base):
        return jsonify({"types": []})

    types = []
    for type_name in sorted(os.listdir(base)):
        type_path = os.path.join(base, type_name)
        if type_name.startswith(".") or not os.path.isdir(type_path):
            continue
        variants = []
        for var_name in sorted(os.listdir(type_path)):
            var_path = os.path.join(type_path, var_name)
            if var_name.startswith(".") or not os.path.isdir(var_path):
                continue
            has_pdf = any(
                f.lower().endswith(".pdf")
                for dirpath, _, files in os.walk(var_path)
                for f in files
            )
            variants.append({"name": var_name, "empty": not has_pdf})
        types.append({"name": type_name, "variants": variants})
    return jsonify({"types": types})


@app.route("/api/products/files", methods=["GET"])
def products_files():
    """Список файлов конкретного варианта изделия."""
    ptype = request.args.get("type", "")
    variant = request.args.get("variant", "")
    if not ptype or not variant:
        return jsonify({"error": "Параметры type и variant обязательны"}), 400

    var_path = _safe_products_path(ptype, variant)
    if not var_path or not os.path.isdir(var_path):
        return jsonify({"error": "Вариант не найден"}), 404

    def scan_dir(dirpath, rel_prefix=""):
        entries = []
        items = sorted(os.listdir(dirpath))
        dirs = [i for i in items if not i.startswith(".") and os.path.isdir(os.path.join(dirpath, i))]
        files = [i for i in items if not i.startswith(".") and i.lower().endswith(".pdf") and os.path.isfile(os.path.join(dirpath, i))]
        for d in dirs:
            children = scan_dir(os.path.join(dirpath, d), rel_prefix + d + "/")
            if children:
                entries.append({"path": rel_prefix + d, "name": d, "type": "dir", "children": children})
        for f in files:
            entries.append({"path": rel_prefix + f, "name": f, "type": "file"})
        return entries

    tree = scan_dir(var_path)
    return jsonify({"variant": variant, "tree": tree})


@app.route("/api/products/pdf", methods=["GET"])
def products_pdf():
    """Отдаёт PDF-файл изделия."""
    ptype = request.args.get("type", "")
    variant = request.args.get("variant", "")
    fpath = request.args.get("path", "")
    if not ptype or not variant or not fpath:
        return jsonify({"error": "Параметры type, variant и path обязательны"}), 400

    if not fpath.lower().endswith(".pdf"):
        return jsonify({"error": "Допустимы только PDF-файлы"}), 400

    full = _safe_products_path(ptype, variant, fpath)
    if not full or not os.path.isfile(full):
        return jsonify({"error": "Файл не найден"}), 404

    return send_file(full, mimetype="application/pdf")


@app.route("/api/products/results", methods=["GET"])
def products_results():
    """Возвращает кешированные результаты анализа для варианта изделия."""
    ptype = request.args.get("type", "")
    variant = request.args.get("variant", "")
    if not ptype or not variant:
        return jsonify({"error": "Параметры type и variant обязательны"}), 400

    var_path = _safe_products_path(ptype, variant)
    if not var_path or not os.path.isdir(var_path):
        return jsonify({"error": "Вариант не найден"}), 404

    cache = _load_products_cache(var_path)
    return jsonify({"results": cache})


def analyze_drawing_from_disk(pdf_path):
    """Анализ чертежа с диска — замечания и оптимизации (как в основном «Анализ чертежа»)."""
    use_stub = os.getenv("USE_STUB", "false").lower() == "true"
    use_claude_code = os.getenv("USE_CLAUDE_CODE", "false").lower() == "true"

    if use_stub:
        return {
            "summary": "Тестовый анализ чертежа.",
            "remarks": [
                {"id": 1, "type": "замечание", "category": "допуски", "priority": "средний",
                 "title": "Тестовое замечание", "description": "Описание.", "suggestion": "Рекомендация.",
                 "x": 50, "y": 50, "w": 10, "h": 10}
            ]
        }

    if use_claude_code:
        prompt = (
            f"Прочитай PDF-файл чертежа детали: {pdf_path}\n\n"
            f"Проведи детальный технический анализ:\n"
            f"1. Найди замечания (ошибки, нарушения ГОСТ, отсутствующие данные)\n"
            f"2. Найди оптимизации (предложения по улучшению технологичности)\n"
            f"Для каждого пункта укажи приблизительные координаты на листе (x,y,w,h в процентах 0-100).\n"
            f"Верни ТОЛЬКО JSON-объект по формату из системного промпта."
        )
        env = os.environ.copy()
        env.pop("CLAUDECODE", None)
        result = subprocess.run(
            [CLAUDE_BIN, "-p", prompt, "--allowedTools", "Read", "--output-format", "json",
             "--system-prompt", DRAWING_SYSTEM_PROMPT, "--model", "sonnet", "--no-session-persistence"],
            capture_output=True, text=True, timeout=300, env=env,
        )
        if result.returncode != 0:
            raise RuntimeError(f"Claude Code ошибка: {result.stderr[:500]}")
        data = json.loads(result.stdout)
        if data.get("is_error"):
            raise RuntimeError(data.get("result", "ошибка"))
        response_text = data["result"].strip()
    else:
        api_key = os.getenv("ANTHROPIC_API_KEY")
        if not api_key:
            raise ValueError("ANTHROPIC_API_KEY не задан")
        with open(pdf_path, "rb") as f:
            pdf_b64 = base64.standard_b64encode(f.read()).decode("utf-8")
        content = [
            {"type": "text", "text": "Проведи детальный технический анализ этого чертежа детали:"},
            {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": pdf_b64}},
            {"type": "text", "text": "Выяви замечания и оптимизации. Верни строго JSON-объект по формату."},
        ]
        client = Anthropic(api_key=api_key)
        message = client.messages.create(
            model="claude-sonnet-4-20250514", max_tokens=4096,
            system=DRAWING_SYSTEM_PROMPT,
            messages=[{"role": "user", "content": content}],
        )
        response_text = message.content[0].text.strip()

    if response_text.startswith("```"):
        lines = response_text.split("\n")
        response_text = "\n".join(lines[1:-1])
    return json.loads(response_text)


@app.route("/api/products/analyze_drawing", methods=["POST"])
def products_analyze_drawing():
    """Анализ чертежа детали из папки изделия."""
    data = request.get_json()
    ptype = data.get("type", "")
    variant = data.get("variant", "")
    filename = data.get("filename", "")

    if not ptype or not variant or not filename:
        return jsonify({"error": "Параметры обязательны"}), 400

    pdf_full = _safe_products_path(ptype, variant, filename)
    if not pdf_full or not os.path.isfile(pdf_full):
        return jsonify({"error": "Файл не найден"}), 404

    try:
        result = analyze_drawing_from_disk(pdf_full)
        # Кешируем
        var_path = _safe_products_path(ptype, variant)
        cache = _load_products_cache(var_path)
        if filename not in cache:
            cache[filename] = {}
        cache[filename]["drawing_analysis"] = result
        _save_products_cache(var_path, cache)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


BOM_PROMPT = """Проанализируй сборочный чертёж и извлеки спецификацию (ведомость деталей).

Верни ТОЛЬКО JSON объект в формате:
{
  "bom": [
    {"деталь": "название детали", "количество": 4},
    {"деталь": "другая деталь", "количество": 2}
  ]
}

Извлеки ВСЕ детали из спецификации/штампа чертежа с их количеством.
Названия деталей пиши так, как они указаны в чертеже.
Верни ТОЛЬКО JSON, без пояснений."""


def extract_bom_from_assembly(pdf_path):
    """Извлекает спецификацию (BOM) из сборочного чертежа."""
    fname = os.path.basename(pdf_path)
    use_stub = os.getenv("USE_STUB", "false").lower() == "true"
    use_claude_code = os.getenv("USE_CLAUDE_CODE", "false").lower() == "true"

    if use_stub:
        return []

    if use_claude_code:
        env = os.environ.copy()
        env.pop("CLAUDECODE", None)
        prompt = f"Прочитай PDF файл:\n1. Сборочный чертёж: {pdf_path}\n\n{BOM_PROMPT}"
        result = subprocess.run(
            [CLAUDE_BIN, "-p", prompt, "--allowedTools", "Read",
             "--output-format", "json", "--model", "sonnet", "--no-session-persistence"],
            capture_output=True, text=True, timeout=300, env=env,
        )
        if result.returncode != 0:
            return []
        data = json.loads(result.stdout)
        response_text = data.get("result", "").strip()
    else:
        api_key = os.getenv("ANTHROPIC_API_KEY")
        if not api_key:
            return []
        with open(pdf_path, "rb") as f:
            pdf_b64 = base64.standard_b64encode(f.read()).decode("utf-8")
        client = Anthropic(api_key=api_key)
        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=2048,
            messages=[{"role": "user", "content": [
                {"type": "text", "text": "Сборочный чертёж:"},
                {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": pdf_b64}},
                {"type": "text", "text": BOM_PROMPT},
            ]}],
        )
        response_text = message.content[0].text.strip()

    if response_text.startswith("```"):
        lines = response_text.split("\n")
        response_text = "\n".join(lines[1:-1])

    parsed = json.loads(response_text)
    return parsed.get("bom", [])


@app.route("/api/products/extract_bom", methods=["POST"])
def products_extract_bom():
    """Извлекает спецификацию из сборочного чертежа и сохраняет количества в кеш."""
    data = request.get_json()
    ptype = data.get("type", "")
    variant = data.get("variant", "")
    assembly_file = data.get("assembly_file", "")

    if not ptype or not variant or not assembly_file:
        return jsonify({"error": "Параметры type, variant и assembly_file обязательны"}), 400

    pdf_full = _safe_products_path(ptype, variant, assembly_file)
    if not pdf_full or not os.path.isfile(pdf_full):
        return jsonify({"error": "Сборочный чертёж не найден"}), 404

    try:
        print(f"[BOM] Извлечение спецификации из {assembly_file}", flush=True)
        bom = extract_bom_from_assembly(pdf_full)
        print(f"[BOM] Найдено {len(bom)} позиций", flush=True)

        # Сохраняем количества в кеш
        var_path = _safe_products_path(ptype, variant)
        cache = _load_products_cache(var_path)
        cache["_bom"] = bom
        _save_products_cache(var_path, cache)

        return jsonify({"bom": bom})
    except Exception as e:
        print(f"[BOM] Ошибка: {e}", flush=True)
        return jsonify({"error": str(e)}), 500


@app.route("/api/products/save_qty", methods=["POST"])
def products_save_qty():
    """Сохраняет количество деталей в кеш."""
    data = request.get_json()
    if not data:
        return jsonify({"error": "Тело запроса пустое"}), 400

    ptype = data.get("type", "")
    variant = data.get("variant", "")
    filename = data.get("filename", "")
    qty = data.get("qty", 1)

    var_path = _safe_products_path(ptype, variant)
    if not var_path or not os.path.isdir(var_path):
        return jsonify({"error": "Вариант не найден"}), 404

    cache = _load_products_cache(var_path)
    if filename in cache:
        cache[filename]["qty"] = max(1, int(qty))
    else:
        cache[filename] = {"qty": max(1, int(qty))}
    _save_products_cache(var_path, cache)
    return jsonify({"ok": True})


@app.route("/api/products/analyze_part", methods=["POST"])
def products_analyze_part():
    """Анализирует один PDF из папки изделия и кеширует результат."""
    data = request.get_json()
    if not data:
        return jsonify({"error": "Тело запроса пустое"}), 400

    ptype = data.get("type", "")
    variant = data.get("variant", "")
    filename = data.get("filename", "")
    if not ptype or not variant or not filename:
        return jsonify({"error": "Параметры type, variant и filename обязательны"}), 400

    if not filename.lower().endswith(".pdf"):
        return jsonify({"error": "Допустимы только PDF-файлы"}), 400

    var_path = _safe_products_path(ptype, variant)
    if not var_path or not os.path.isdir(var_path):
        return jsonify({"error": "Вариант не найден"}), 404

    pdf_full = _safe_products_path(ptype, variant, filename)
    if not pdf_full or not os.path.isfile(pdf_full):
        return jsonify({"error": "Файл не найден"}), 404

    try:
        alog(filename, f"Запрос анализа: {variant} / {filename}")
        result = analyze_pdf_from_disk(pdf_full, batch_size=1)
        operations = result.get("операции", [])
        route = result.get("маршрут", {})

        # Кешируем результат
        cache = _load_products_cache(var_path)
        cache[filename] = {
            "operations": operations,
            "route": route,
            "analyzed_at": datetime.now().isoformat()
        }
        _save_products_cache(var_path, cache)
        alog(filename, f"Результат сохранён в кеш ({len(operations)} операций)", "success")

        return jsonify({"filename": filename, "operations": operations, "route": route})

    except Exception as e:
        alog(filename, f"Ошибка: {str(e)}", "error")
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    init_db()
    app.run(debug=True, port=5050, use_reloader=False, threaded=True)
